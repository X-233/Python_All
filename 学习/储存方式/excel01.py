# 1、爬虫速度不要太快，不要给对方服务器造成太大压力
# 2、爬虫不要伪造VIP，绕过对方身份验证，可以买一个VIP做自动化
# 3、公民个人信息不要去碰
import openpyxl as xl

# 1. 创建一个工作对象
work_book = xl.Workbook()

# 2. 创建表对象, 关闭表格程序, work_book.active 默认表对象
sheet_1 = work_book.create_sheet('表一')
print(sheet_1)
# 数据填充
# 2.1: 单个操作
sheet_1['C1'] = 'python'
sheet_1['A3'] = 'python'
# 2.2: 单元格操作
sheet_1.cell(row=1, column=1, value='java')
sheet_1.cell(row=2, column=2).value = 'java2'
# 2.3: 对象调用,append
data_1 = [1, 2, 3, 4, 5]
data_2 = (1, 2, 3, 4, 5)
sheet_1.append(data_1)
sheet_1.append(data_2)

# 3. 保存工作簿对象
work_book.save('示例1.xlsx')
